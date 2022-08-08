# Dependencies
import tkinter as tk
from tkinter import ttk
import openpyxl as xl

# Dictionary of Company Number to Company Name
company_dict = {
    '1' : 'CVC Construction Corp.',
    '2' : 'CVC Holding Corp ',
    '3' : 'CVC Equipment LLC',
    '6' : 'Cedar Valley Concrete Corp of NV',
    '8' : 'Concrete Value Corp',
    '9' : 'CVC Commercial Corp',
    '11': 'Concrete Value Corp of Nevada',
    '12': 'Hilo Erectors',
    '30': 'DVC Concrete Corp',
    '40': 'CRC Trucking LLC'
}


# Dictionary of Vendor Type
vendor_type_list = [
    'Atty. / Consultant. (1099 required)',
    'Corp. / Inc. (no 1099 required)',
    'Gov. (no 1099 required)',
    'Emp. Adv. or Reimb. (no 1099 required)',
]

# List of all the People who are authorized to fill out the form
# TODO : Add names of all the people who are 
request_by_list = [
    'Nathan Esch'
]

print(company_dict.values())

# Class for the GUI
class App(tk.Tk):

    

    # Constructor
    def __init__(self):
        super().__init__()
        self.title('Vendor Setup Form Filler')
        self.geometry('300x500')

        # Icon for the GUI
        icon = tk.PhotoImage(file='Resources/CVC-logo.png')
        self.iconphoto(False, icon)

        self.resizable(False, False)

        # Set an instance of the FormFiller class
        self.form = FormFiller()

        
        # Labels and Entries for the form and set them to the grid.
        # Selection of the Company needing the form 
        self.lbl_company = ttk.Label(self, text='Company')
        self.lbl_company.grid(row=0, column=0, sticky=tk.W)
        self.company_var = ttk.Combobox(self, values=list(company_dict.values()))
        self.company_var.grid(row=0, column=1, sticky=tk.W)

        # Vendor Details
        # Vendor Name
        self.lbl_vendor_name = ttk.Label(self, text='Vendor Name')
        self.lbl_vendor_name.grid(row=1, column=0, sticky=tk.W)
        self.vendor_name_var = ttk.Entry(self)
        self.vendor_name_var.grid(row=1, column=1, sticky=tk.W)

        # Vendor Street Address
        self.lbl_vendor_street = ttk.Label(self, text='Vendor Street')
        self.lbl_vendor_street.grid(row=2, column=0, sticky=tk.W)
        self.vendor_street_var = ttk.Entry(self)
        self.vendor_street_var.grid(row=2, column=1, sticky=tk.W)

        # Vendor City, State, Zip
        self.lbl_vendor_city = ttk.Label(self, text='Vendor City, State, Zip')
        self.lbl_vendor_city.grid(row=3, column=0, sticky=tk.W)
        self.vendor_city_var = ttk.Entry(self)
        self.vendor_city_var.grid(row=3, column=1, sticky=tk.W)

        # Vendor Phone/Fax Number
        self.lbl_vendor_phone = ttk.Label(self, text='Vendor Phone/Fax')
        self.lbl_vendor_phone.grid(row=4, column=0, sticky=tk.W)
        self.vendor_phone_var = ttk.Entry(self)
        self.vendor_phone_var.grid(row=4, column=1, sticky=tk.W)

        # Vendor Tax ID
        self.lbl_vendor_taxid = ttk.Label(self, text='Vendor Tax ID')
        self.lbl_vendor_taxid.grid(row=5, column=0, sticky=tk.W)
        self.vendor_taxid_var = ttk.Entry(self)
        self.vendor_taxid_var.grid(row=5, column=1, sticky=tk.W)

        # Vendor Number
        self.lbl_vendor_number = ttk.Label(self, text='Vendor Number')
        self.lbl_vendor_number.grid(row=6, column=0, sticky=tk.W)
        self.vendor_number_var = ttk.Entry(self)
        self.vendor_number_var.grid(row=6, column=1, sticky=tk.W)

        # Financial Information
        # Amount of the Vendor's Invoice???
        self.lbl_invoice_amount = ttk.Label(self, text='Invoice Amount')
        self.lbl_invoice_amount.grid(row=7, column=0, sticky=tk.W)
        self.invoice_amount_var = ttk.Entry(self)
        self.invoice_amount_var.grid(row=7, column=1, sticky=tk.W)

        # Job Code
        self.lbl_job_code = ttk.Label(self, text='Job Code')
        self.lbl_job_code.grid(row=8, column=0, sticky=tk.W)
        self.job_code_var = ttk.Entry(self)
        self.job_code_var.grid(row=8, column=1, sticky=tk.W)

        # Invoice Attached
        self.lbl_invoice_attached = ttk.Label(self, text='Invoice Attached?')
        self.lbl_invoice_attached.grid(row=9, column=0, sticky=tk.W)
        self.invoice_attached_var = ttk.Combobox(self, values=['Yes', 'No'])
        self.invoice_attached_var.grid(row=9, column=1, sticky=tk.W)

        # Invoice Number
        self.lbl_invoice_number = ttk.Label(self, text='Invoice Number')
        self.lbl_invoice_number.grid(row=10, column=0, sticky=tk.W)
        self.invoice_number_var = ttk.Entry(self)
        self.invoice_number_var.grid(row=10, column=1, sticky=tk.W)

        # Invoice Amount
        self.lbl_invoice_amount = ttk.Label(self, text='Amount')
        self.lbl_invoice_amount.grid(row=11, column=0, sticky=tk.W)
        self.invoice_amount_var = ttk.Entry(self)
        self.invoice_amount_var.grid(row=11, column=1, sticky=tk.W)

        # Date Needed
        self.lbl_date_needed = ttk.Label(self, text='Date Needed')
        self.lbl_date_needed.grid(row=12, column=0, sticky=tk.W)
        self.date_needed_var = ttk.Entry(self)
        self.date_needed_var.grid(row=12, column=1, sticky=tk.W)

        # Date
        self.lbl_date = ttk.Label(self, text='Date')
        self.lbl_date.grid(row=13, column=0, sticky=tk.W)
        self.date_var = ttk.Entry(self)
        self.date_var.grid(row=13, column=1, sticky=tk.W)

        # Brief Description
        self.lbl_brief_description = ttk.Label(self, text='Brief Description')
        self.lbl_brief_description.grid(row=14, column=0, sticky=tk.W)
        self.brief_description_var = ttk.Entry(self)
        self.brief_description_var.grid(row=14, column=1, sticky=tk.W)
        
        # Special Instructions
        self.lbl_special_instructions = ttk.Label(self, text='Special Instructions')
        self.lbl_special_instructions.grid(row=15, column=0, sticky=tk.W)
        self.special_instructions_var = ttk.Entry(self)
        self.special_instructions_var.grid(row=15, column=1, sticky=tk.W)
        
        # Vendor Type
        self.lbl_vendor_type = ttk.Label(self, text='Vendor Type')
        self.lbl_vendor_type.grid(row=16, column=0, sticky=tk.W)
        self.vendor_type_var = ttk.Combobox(self, values= vendor_type_list)
        self.vendor_type_var.grid(row=16, column=1, sticky=tk.W)

        # Requested By
        self.lbl_requested_by = ttk.Label(self, text='Requested By')
        self.lbl_requested_by.grid(row=17, column=0, sticky=tk.W)
        self.requested_by_var = ttk.Combobox(self, values=request_by_list)
        self.requested_by_var.grid(row=17, column=1, sticky=tk.W)


        # Create Form Button
        self.btn_create_form = ttk.Button(self, text='Create Form', command=self.create_form)
        self.btn_create_form.grid(row=18, column=0, sticky=tk.W)
        # self.btn_create_form.config(state=tk.DISABLED) # What does this do?

        # Close Button
        self.btn_close = ttk.Button(self, text='Close', command=self.destroy)
        self.btn_close.grid(row=18, column=1, sticky=tk.W)


    # Create Form Function
    def create_form(self):
        # Get values from form
        company = self.company_var.current()
        vendor_name = self.vendor_name_var.get()
        vendor_street = self.vendor_street_var.get()
        vendor_city = self.vendor_city_var.get()
        vendor_phone = self.vendor_phone_var.get()
        vendor_taxid = self.vendor_taxid_var.get()
        vendor_number = self.vendor_number_var.get()
        invoice_amount = self.invoice_amount_var.get()
        job_code = self.job_code_var.get()
        invoice_attached = self.invoice_attached_var.get()
        invoice_number = self.invoice_number_var.get()
        invoice_amount = self.invoice_amount_var.get()
        date_needed = self.date_needed_var.get()
        date = self.date_var.get()
        brief_description = self.brief_description_var.get()
        special_instructions = self.special_instructions_var.get()
        vendor_type = self.vendor_type_var.current()
        requested_by = self.requested_by_var.get()

        # Create form
        self.form.fill_form([company, vendor_name, vendor_street, vendor_city,
                             vendor_phone, vendor_taxid, vendor_number,
                             job_code, invoice_attached, invoice_number,
                             invoice_amount, date_needed, date, brief_description,
                             special_instructions, vendor_type, requested_by])

        # Clear form
        self.vendor_name_var.delete(0, tk.END)
        self.vendor_street_var.delete(0, tk.END)
        self.vendor_city_var.delete(0, tk.END)
        self.vendor_phone_var.delete(0, tk.END)
        self.vendor_taxid_var.delete(0, tk.END)
        self.vendor_number_var.delete(0, tk.END)
        self.invoice_amount_var.delete(0, tk.END)
        self.job_code_var.delete(0, tk.END)
        self.invoice_attached_var.delete(0, tk.END)
        self.invoice_number_var.delete(0, tk.END)
        self.invoice_amount_var.delete(0, tk.END)
        self.date_needed_var.delete(0, tk.END)
        self.date_var.delete(0, tk.END)
        self.brief_description_var.delete(0, tk.END)
        self.special_instructions_var.delete(0, tk.END)
        self.vendor_type_var.delete(0, tk.END)
        self.requested_by_var.delete(0, tk.END)


# Class to fill out forms
class FormFiller:
    # Constructor
    def __init__(self):
        self.wb = xl.load_workbook('Resources/Template.xlsx')
        self.ws = self.wb.active

    # Function to fill out form
    # Data format:
    # [company, vendor_name, vendor_street, vendor_city,
    #  vendor_phone, vendor_taxid, vendor_number,
    #  job_code, invoice_attached, invoice_number,
    #  invoice_amount, date_needed, date, brief_description,
    #  special_instructions, vendor_type, requested_by]
    def fill_form(self, data):
        # Debugging
        print(data)
        
        # Set Company anbd Company No
        # Range of company numbers = [c9:c18]
        self.ws['C' + str(data[0] + 9)] = 'X'
        self.ws['K23'] = self.ws['D'+str(data[0] + 9)].value

        # Set Vendor Name, Street Address, City, State, and Zip
        self.ws['G21'] = data[1]
        self.ws['G22'] = data[2]
        self.ws['G23'] = data[3]

        # Set Vendor Phone     
        self.ws['G26'] = data[4]

        # Set Vendor Tax ID
        self.ws['G29'] = data[5]

        # Set Vendor Number
        self.ws['K26'] = data[6]

        # Set Job Code
        self.ws['J30'] = data[7]

        # Set Invoice Attached
        self.ws['G39'] = data[8]

        # Set Invoice Number
        self.ws['J36'] = data[9]

        # Set Invoice Amount
        self.ws['G31'] = data[10]
        self.ws['K30'] = data[10]
        self.ws['K36'] = data[10]
        
        # TODO: Set multiple invoices
        # Grand Total
        self.ws['K39'] = data[10]


        # Set Date Needed
        self.ws['G34'] = data[11]

        # Set Date
        self.ws['G36'] = data[12]

        # Set Brief Description
        self.ws['G43'] = data[13]


        # Set Special Instructions
        self.ws['G47'] = data[14]

        # Set Vendor Type
        self.ws['C' + str(data[15] + 60)] = 'X'

        # Set Requested By
        self.ws['G41'] = data[16]
        

        # Need to find out what kind of name is going to be 
        self.wb.save('Test1.xlsx')
        self.wb.close()




# Main
if __name__ == '__main__':
    app = App()
    app.mainloop()


